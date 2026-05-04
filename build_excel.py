"""Build NZ 2026 Trip Planner Excel file."""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
ALPINE   = "0A5C7E"
LAKE     = "0BA3B3"
SL_GREEN = "2A9D8F"
JA_PINK  = "C13A74"
GOLD     = "F4A261"
LIGHT_BG = "EAF4F6"
WHITE    = "FFFFFF"
INK      = "0E2230"
INK_SOFT = "4A5868"
WARN_BG  = "FEF3E2"
WARN_BDR = "F4A261"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(color="CCCCCC", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(color=WHITE, size=11, bold=True):
    return Font(name="Calibri", color=color, size=size, bold=bold)

def cell_font(color=INK, size=10, bold=False):
    return Font(name="Calibri", color=color, size=size, bold=bold)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22

# Title
ws1.merge_cells("A1:D1")
t = ws1["A1"]
t.value = "🥝  NZ 2026 · genz bebe Trip Summary"
t.font = Font(name="Calibri", color=WHITE, size=16, bold=True)
t.fill = fill(ALPINE)
t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 40

# Sub-header
ws1.merge_cells("A2:D2")
s = ws1["A2"]
s.value = "May 16 – 30, 2026  ·  South Island, New Zealand"
s.font = Font(name="Calibri", color=WHITE, size=11)
s.fill = fill(LAKE)
s.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 24

# Blank row
ws1.row_dimensions[3].height = 8

# ── TRIP STATS ────────────────────────────────────────────────────────────────
stats_header = ws1["A4"]
stats_header.value = "TRIP STATS"
stats_header.font = hdr_font(color=ALPINE, size=10)
stats_header.fill = fill("DCF0F5")
stats_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.merge_cells("A4:D4")
ws1.row_dimensions[4].height = 20

stats = [
    ("Total Days on Island", 15, "", ""),
    ("Total Nights", 14, "", ""),
    ("Approx Drive Distance", "~1,095 km", "", ""),
    ("Departure (CHC)", "May 16, 2026", "", ""),
    ("Return (WKA→CHC→home)", "May 30, 2026", "", ""),
    ("Phases / Cities", 5, "", ""),
]
for i, (label, val, _, _) in enumerate(stats, start=5):
    ws1[f"A{i}"].value = label
    ws1[f"A{i}"].font = cell_font(bold=True)
    ws1[f"A{i}"].fill = fill("F5FBFC")
    ws1[f"A{i}"].border = border("E0E0E0")
    ws1[f"B{i}"].value = val
    ws1[f"B{i}"].font = cell_font()
    ws1[f"B{i}"].fill = fill(WHITE)
    ws1[f"B{i}"].border = border("E0E0E0")
    ws1.merge_cells(f"C{i}:D{i}")
    ws1.row_dimensions[i].height = 18

# Blank
ws1.row_dimensions[11].height = 10

# ── COST SUMMARY ─────────────────────────────────────────────────────────────
ws1.merge_cells("A12:D12")
ch = ws1["A12"]
ch.value = "COST SUMMARY (USD)"
ch.font = hdr_font(color=ALPINE, size=10)
ch.fill = fill("DCF0F5")
ch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.row_dimensions[12].height = 20

cost_hdr_row = 13
for col, label in zip("ABCD", ["Category", "Stephen & Lindsay", "Jared & Ariel", "Combined"]):
    c = ws1[f"{col}{cost_hdr_row}"]
    c.value = label
    c.font = hdr_font(size=10)
    c.fill = fill(ALPINE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border(WHITE)
ws1.row_dimensions[cost_hdr_row].height = 20

cost_rows = [
    ("Lodging",      3247, 3038, 6285),
    ("Activities",    758,  825, 1583),
    ("Car Rental",    769,  385, 1154),
    ("TOTAL",        4774, 4248, 9022),
]
for i, (cat, sl, ja, combo) in enumerate(cost_rows, start=14):
    is_total = cat == "TOTAL"
    bg = "E8F6F8" if not is_total else "D4EEF2"
    ws1[f"A{i}"].value = cat
    ws1[f"A{i}"].font = cell_font(bold=is_total)
    ws1[f"A{i}"].fill = fill(bg)
    ws1[f"A{i}"].border = border("C8E8EE")
    for col, val in zip("BCD", [sl, ja, combo]):
        c = ws1[f"{col}{i}"]
        c.value = val
        c.number_format = '"$"#,##0'
        c.font = cell_font(bold=is_total, color=SL_GREEN if col == "B" else (JA_PINK if col == "C" else ALPINE))
        c.fill = fill(bg)
        c.border = border("C8E8EE")
        c.alignment = Alignment(horizontal="right")
    ws1.row_dimensions[i].height = 18

# Blank
ws1.row_dimensions[18].height = 10

# ── BALANCE ───────────────────────────────────────────────────────────────────
ws1.merge_cells("A19:D19")
bal = ws1["A19"]
bal.value = "NET BALANCE:  J&A owes S&L  $191.48"
bal.font = Font(name="Calibri", color=WHITE, size=12, bold=True)
bal.fill = fill(SL_GREEN)
bal.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[19].height = 28

ws1.merge_cells("A20:D20")
note = ws1["A20"]
note.value = ("J&A → S&L:  $1,027.71 (Wanaka) + $384.67 (car)  =  $1,412.38"
              "        |        "
              "S&L → J&A:  $768.20 (QT) + $217.70 (Te Anau) + $235.00 (Milford)  =  $1,220.90")
note.font = Font(name="Calibri", color=INK_SOFT, size=9)
note.fill = fill("F0FAF5")
note.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[20].height = 30

# Blank
ws1.row_dimensions[21].height = 10

# ── UNRESOLVED ────────────────────────────────────────────────────────────────
ws1.merge_cells("A22:D22")
uh = ws1["A22"]
uh.value = "⚠  STILL TO RESOLVE"
uh.font = Font(name="Calibri", color="A06A34", size=10, bold=True)
uh.fill = fill(WARN_BG)
uh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.row_dimensions[22].height = 20

unresolved = [
    "Skyline Gondola + Luge (~$240) — pay at venue, split TBD",
    "Gibbston wineries (May 23) — each couple covers their own tastings & food",
]
for i, item in enumerate(unresolved, start=23):
    ws1.merge_cells(f"A{i}:D{i}")
    c = ws1[f"A{i}"]
    c.value = f"  • {item}"
    c.font = cell_font(color="7A5020")
    c.fill = fill(WARN_BG)
    c.border = border(WARN_BDR)
    ws1.row_dimensions[i].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — ITINERARY
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Itinerary")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 8   # Day #
ws2.column_dimensions["B"].width = 14  # Date
ws2.column_dimensions["C"].width = 16  # Location
ws2.column_dimensions["D"].width = 30  # Accommodation
ws2.column_dimensions["E"].width = 40  # Activities / Notes
ws2.column_dimensions["F"].width = 14  # Phase

# Title
ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value = "🗓  NZ 2026 · Day-by-Day Itinerary"
t2.font = Font(name="Calibri", color=WHITE, size=14, bold=True)
t2.fill = fill(ALPINE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

# Column headers
headers = ["Day", "Date", "Location", "Accommodation", "Activities / Notes", "Phase"]
hdr_row = 2
for col_idx, h in enumerate(headers, start=1):
    c = ws2.cell(row=hdr_row, column=col_idx, value=h)
    c.font = hdr_font(size=10)
    c.fill = fill(LAKE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border(WHITE)
ws2.row_dimensions[hdr_row].height = 22

# Phase colours
PHASE_COLORS = {
    "Solo": "F5E6CC",
    "Christchurch": "D6EEF5",
    "Tekapo": "D6EEF5",
    "Queenstown": "DDF0EC",
    "Te Anau": "E8F4D6",
    "Wanaka": "F0E8F4",
    "Travel": "F0F0F0",
}

itinerary = [
    # day, date, location, accommodation, activities, phase
    (1,  "Fri May 15", "Travel Day",    "—",                               "Depart US → Auckland",                                          "Travel"),
    (2,  "Sat May 16", "Christchurch",  "CHC CBD Townhouse",                "Arrive CHC · explore city · Hagley Park · botanical gardens",    "Christchurch"),
    (3,  "Sun May 17", "Christchurch",  "CHC CBD Townhouse",                "Akaroa Dolphin Cruise · Banks Peninsula drive",                  "Christchurch"),
    (4,  "Mon May 18", "Lake Tekapo",   "Peppers Bluewater Resort",         "Drive CHC → Tekapo · Lake Tekapo · Church of the Good Shepherd", "Tekapo"),
    (5,  "Tue May 19", "Lake Tekapo",   "Peppers Bluewater Resort",         "Dark Sky Stargazing · explore Tekapo",                           "Tekapo"),
    (6,  "Wed May 20", "Lake Tekapo",   "Peppers Bluewater Resort",         "Drive to Queenstown via Lindis Pass",                            "Tekapo"),
    (7,  "Thu May 21", "Queenstown",    "Queenstown AirBnb (4 couples)",    "J&A arrive · regroup · explore Queenstown waterfront",           "Queenstown"),
    (8,  "Fri May 22", "Queenstown",    "Queenstown AirBnb",                "Skyline Gondola + Luge · Queenstown gardens",                    "Queenstown"),
    (9,  "Sat May 23", "Queenstown",    "Queenstown AirBnb",                "Gibbston Valley wineries · Arrow River trail",                   "Queenstown"),
    (10, "Sun May 24", "Te Anau",       "Te Anau AirBnb",                   "Drive QT → Te Anau · Fiordland NP",                              "Te Anau"),
    (11, "Mon May 25", "Te Anau",       "Te Anau AirBnb",                   "Milford Sound Boutique Cruise (booking #346437283)",             "Te Anau"),
    (12, "Tue May 26", "Wanaka",        "The Lookout · Lake Views",         "Drive Te Anau → Wanaka via Queenstown",                          "Wanaka"),
    (13, "Wed May 27", "Wanaka",        "The Lookout · Lake Views",         "Wanaka Tree · Roys Peak track · town explore",                   "Wanaka"),
    (14, "Thu May 28", "Wanaka",        "The Lookout · Lake Views",         "Puzzling World · Glendhu Bay · local hikes",                     "Wanaka"),
    (15, "Fri May 29", "Wanaka",        "The Lookout · Lake Views",         "Final Wanaka day · farewell dinner",                             "Wanaka"),
    (16, "Sat May 30", "Travel Day",    "—",                                "Drive WKA → CHC · fly home",                                     "Travel"),
]

for row_idx, (day, date, loc, accom, activities, phase) in enumerate(itinerary, start=3):
    bg = PHASE_COLORS.get(phase, WHITE)
    row_data = [day, date, loc, accom, activities, phase]
    for col_idx, val in enumerate(row_data, start=1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.font = cell_font(bold=(col_idx == 1))
        c.fill = fill(bg)
        c.border = border("D8D8D8")
        c.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))
    ws2.row_dimensions[row_idx].height = 20

# freeze header rows
ws2.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — EXPENSES
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Expenses")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 10  # Paid by
ws3.column_dimensions["B"].width = 12  # Date
ws3.column_dimensions["C"].width = 32  # Description
ws3.column_dimensions["D"].width = 14  # Amount (USD)
ws3.column_dimensions["E"].width = 28  # Split notes
ws3.column_dimensions["F"].width = 16  # Phase

# Title
ws3.merge_cells("A1:F1")
t3 = ws3["A1"]
t3.value = "💳  NZ 2026 · Expense Splits"
t3.font = Font(name="Calibri", color=WHITE, size=14, bold=True)
t3.fill = fill(ALPINE)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

# Column headers
exp_headers = ["Paid By", "Date", "Description", "Amount (USD)", "Split", "Phase"]
for col_idx, h in enumerate(exp_headers, start=1):
    c = ws3.cell(row=2, column=col_idx, value=h)
    c.font = hdr_font(size=10)
    c.fill = fill(LAKE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border(WHITE)
ws3.row_dimensions[2].height = 22

expenses = [
    # paid_by, date, description, amount, split_note, phase
    ("S&L", "May 16–20", "Car rental · 5 solo days (~NZD $658.67 @ 0.584)",  384.66, "S&L only — no split",           "Car"),
    ("S&L", "May 21–30", "Car rental · 10 shared days (~NZD $1,317.35 @ 0.584)", 769.33, "50/50 → J&A owes S&L $384.67", "Car"),
    ("S&L", "May 16",    "CHC CBD Townhouse",                           250.00, "S&L only",                      "Christchurch"),
    ("S&L", "May 17",    "CHC CBD Townhouse",                           250.00, "S&L only",                      "Christchurch"),
    ("S&L", "May 17",    "Akaroa Dolphin Cruise",                       148.20, "S&L only",                      "Christchurch"),
    ("S&L", "May 18",    "Peppers Bluewater Resort (Tekapo)",            244.33, "S&L only",                      "Tekapo"),
    ("S&L", "May 19",    "Peppers Bluewater Resort (Tekapo)",            244.33, "S&L only",                      "Tekapo"),
    ("S&L", "May 19",    "Dark Sky Stargazing",                         254.65, "S&L only",                      "Tekapo"),
    ("S&L", "May 20",    "Peppers Bluewater Resort (Tekapo)",            244.33, "S&L only",                      "Tekapo"),
    ("J&A", "May 19",    "Queenstown AirBnb",                           512.13, "J&A only",                      "Queenstown"),
    ("J&A", "May 20",    "Queenstown AirBnb",                           512.13, "J&A only",                      "Queenstown"),
    ("J&A", "May 21",    "Queenstown AirBnb",                           512.13, "50/50 → S&L owes J&A $256.07",  "Queenstown"),
    ("J&A", "May 22",    "Queenstown AirBnb",                           512.13, "50/50 → S&L owes J&A $256.07",  "Queenstown"),
    ("J&A", "May 23",    "Queenstown AirBnb",                           512.13, "50/50 → S&L owes J&A $256.07",  "Queenstown"),
    ("J&A", "May 24",    "Te Anau AirBnb",                              217.70, "50/50 → S&L owes J&A $108.85",  "Te Anau"),
    ("J&A", "May 25",    "Te Anau AirBnb",                              217.70, "50/50 → S&L owes J&A $108.85",  "Te Anau"),
    ("J&A", "May 25",    "Milford Sound Boutique Cruise (#346437283)",   470.00, "50/50 → S&L owes J&A $235.00",  "Te Anau"),
    ("S&L", "May 26",    "The Lookout · Lake Views (Wanaka)",           513.86, "50/50 → J&A owes S&L $256.93",  "Wanaka"),
    ("S&L", "May 27",    "The Lookout · Lake Views (Wanaka)",           513.86, "50/50 → J&A owes S&L $256.93",  "Wanaka"),
    ("S&L", "May 28",    "The Lookout · Lake Views (Wanaka)",           513.86, "50/50 → J&A owes S&L $256.93",  "Wanaka"),
    ("S&L", "May 29",    "The Lookout · Lake Views (Wanaka)",           513.86, "50/50 → J&A owes S&L $256.93",  "Wanaka"),
]

PHASE_EXP_COLORS = {
    "Car":          "FFF8EE",
    "Christchurch": "D6EEF5",
    "Tekapo":       "D6EEF5",
    "Queenstown":   "DDF0EC",
    "Te Anau":      "E8F4D6",
    "Wanaka":       "F0E8F4",
}

current_phase = None
row_idx = 3
for paid_by, date, desc, amount, split, phase in expenses:
    # Insert phase header row when phase changes
    if phase != current_phase:
        current_phase = phase
        ws3.merge_cells(f"A{row_idx}:F{row_idx}")
        ph = ws3[f"A{row_idx}"]
        ph.value = f"── {phase} ──"
        ph.font = Font(name="Calibri", color=ALPINE, size=9, bold=True)
        ph.fill = fill("E8F4F8")
        ph.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[row_idx].height = 16
        row_idx += 1

    bg = PHASE_EXP_COLORS.get(phase, WHITE)
    paid_color = SL_GREEN if paid_by == "S&L" else JA_PINK

    row_data = [paid_by, date, desc, amount, split, phase]
    for col_idx, val in enumerate(row_data, start=1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        c.fill = fill(bg)
        c.border = border("D8D8D8")
        c.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 5)))
        if col_idx == 1:
            c.font = Font(name="Calibri", color=paid_color, size=10, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 4:
            c.font = cell_font(bold=True)
            c.number_format = '"$"#,##0.00'
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.font = cell_font()
    ws3.row_dimensions[row_idx].height = 18
    row_idx += 1

# Totals row
ws3.row_dimensions[row_idx].height = 6
row_idx += 1
ws3.merge_cells(f"A{row_idx}:C{row_idx}")
ws3[f"A{row_idx}"].value = "TOTAL TRACKED EXPENSES"
ws3[f"A{row_idx}"].font = hdr_font(size=10)
ws3[f"A{row_idx}"].fill = fill(ALPINE)
ws3[f"A{row_idx}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
total_val = sum(e[3] for e in expenses)
ws3[f"D{row_idx}"].value = total_val
ws3[f"D{row_idx}"].number_format = '"$"#,##0.00'
ws3[f"D{row_idx}"].font = hdr_font(size=11)
ws3[f"D{row_idx}"].fill = fill(ALPINE)
ws3[f"D{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
ws3.merge_cells(f"E{row_idx}:F{row_idx}")
ws3[f"E{row_idx}"].fill = fill(ALPINE)
ws3.row_dimensions[row_idx].height = 24

# Balance summary
row_idx += 2
ws3.merge_cells(f"A{row_idx}:F{row_idx}")
bal3 = ws3[f"A{row_idx}"]
bal3.value = "NET BALANCE:  J&A owes S&L  $191.48"
bal3.font = Font(name="Calibri", color=WHITE, size=12, bold=True)
bal3.fill = fill(SL_GREEN)
bal3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[row_idx].height = 28

ws3.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — PACKING
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Packing")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 14
ws4.column_dimensions["D"].width = 28
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 14

# Title
ws4.merge_cells("A1:F1")
t4 = ws4["A1"]
t4.value = "🎒  NZ 2026 · Packing List"
t4.font = Font(name="Calibri", color=WHITE, size=14, bold=True)
t4.fill = fill(ALPINE)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 36

# Two-column layout: S&L left (A-C), J&A right (D-F)
# Sub-headers
for col, label in [("A", "Item"), ("B", "Stephen"), ("C", "Lindsay"),
                   ("D", "Item"), ("E", "Jared"), ("F", "Ariel")]:
    c = ws4[f"{col}2"]
    c.value = label
    c.font = hdr_font(size=10, color=WHITE if col in "ABCDEF" else INK)
    left_fill = SL_GREEN if col in ("A","B","C") else JA_PINK
    c.fill = fill(left_fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border(WHITE)
ws4.row_dimensions[2].height = 22

packing = [
    # (section, item_sl, item_ja)
    ("Clothing", "Hiking pants (2 pairs)", "Hiking pants (2 pairs)"),
    ("Clothing", "Thermal base layer top", "Thermal base layer top"),
    ("Clothing", "Thermal base layer bottom", "Thermal base layer bottom"),
    ("Clothing", "Fleece mid-layer", "Fleece mid-layer"),
    ("Clothing", "Waterproof rain jacket", "Waterproof rain jacket"),
    ("Clothing", "Down puffer jacket", "Down puffer jacket"),
    ("Clothing", "Warm hat & gloves", "Warm hat & gloves"),
    ("Clothing", "Hiking boots (broken in)", "Hiking boots (broken in)"),
    ("Clothing", "Trail running shoes", "Trail running shoes"),
    ("Clothing", "Wool socks (4–5 pairs)", "Wool socks (4–5 pairs)"),
    ("Clothing", "Casual dinner clothes", "Casual dinner clothes"),
    ("Clothing", "Swimwear", "Swimwear"),
    ("Clothing", "Underwear (7 pairs)", "Underwear (7 pairs)"),
    ("", "", ""),
    ("Gear", "Daypack / backpack", "Daypack / backpack"),
    ("Gear", "Trekking poles", "Trekking poles"),
    ("Gear", "Headlamp + batteries", "Headlamp + batteries"),
    ("Gear", "Sunglasses", "Sunglasses"),
    ("Gear", "Sunscreen SPF 50+", "Sunscreen SPF 50+"),
    ("Gear", "Lip balm SPF", "Lip balm SPF"),
    ("Gear", "Blister prevention / moleskin", "Blister prevention / moleskin"),
    ("", "", ""),
    ("Tech", "Passport (valid >6 mo)", "Passport (valid >6 mo)"),
    ("Tech", "Phone + charging cable", "Phone + charging cable"),
    ("Tech", "NZ power adapter (Type I)", "NZ power adapter (Type I)"),
    ("Tech", "Portable battery / power bank", "Portable battery / power bank"),
    ("Tech", "Camera", "Camera"),
    ("Tech", "Download offline maps (NZ)", "Download offline maps (NZ)"),
    ("", "", ""),
    ("Docs & Money", "Travel insurance docs", "Travel insurance docs"),
    ("Docs & Money", "NZD cash (small bills)", "NZD cash (small bills)"),
    ("Docs & Money", "Credit card (no foreign fees)", "Credit card (no foreign fees)"),
    ("Docs & Money", "International driving permit", "International driving permit"),
    ("Docs & Money", "Booking confirmations printed", "Booking confirmations printed"),
]

row_idx = 3
current_section = None
for section, sl_item, ja_item in packing:
    if section and section != current_section:
        current_section = section
        # Section header spanning all 6 cols
        ws4.merge_cells(f"A{row_idx}:F{row_idx}")
        sh = ws4[f"A{row_idx}"]
        sh.value = section.upper()
        sh.font = Font(name="Calibri", color=ALPINE, size=9, bold=True)
        sh.fill = fill("DCF0F5")
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws4.row_dimensions[row_idx].height = 16
        row_idx += 1

    if not sl_item and not ja_item:
        ws4.row_dimensions[row_idx].height = 6
        row_idx += 1
        continue

    for col_idx, val in enumerate([sl_item, "☐", "☐", ja_item, "☐", "☐"], start=1):
        c = ws4.cell(row=row_idx, column=col_idx, value=val)
        is_left = col_idx <= 3
        bg = "F0FAF8" if is_left else "FDF0F6"
        c.fill = fill(bg)
        c.border = border("E0E0E0")
        c.font = cell_font()
        if col_idx in (2, 3, 5, 6):
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Calibri", color=SL_GREEN if is_left else JA_PINK, size=12)
    ws4.row_dimensions[row_idx].height = 18
    row_idx += 1

ws4.freeze_panes = "A3"


# ── SAVE ──────────────────────────────────────────────────────────────────────
output_path = "C:/Users/sbain/claude/Sabbatical/NZ2026_TripPlanner.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
